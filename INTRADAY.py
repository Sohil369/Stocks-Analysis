"""
buzz_intraday_scanner.py

Usage:
    python buzz_intraday_scanner.py

Input:
    - stocks.xlsx: Excel file with tickers in the first column

Output:
    - scores.xlsx: Sorted, color-coded with signals and actions
    - scores.csv: CSV version of results
"""

import os
import time
import math
import pandas as pd
import numpy as np
import yfinance as yf
from pytrends.request import TrendReq
import requests
import google.generativeai as genai
import newsdataapi
from newsdataapi import NewsDataApiClient
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from datetime import datetime, timedelta
import concurrent.futures
import praw

# Optional NewsAPI support
try:
    from newsapi import NewsApiClient
    HAS_NEWSAPI = True
except ImportError:
    HAS_NEWSAPI = False

# ========= CONFIG =========
GEMINI_API_KEY = "AIzaSyATT0mTSy9xHWOSGzT8IMeL-82xyOm4MSo"
NEWSDATA_API_KEY = "pub_1ab52a22ddb54464bc6c94ac1dd7716e"

genai.configure(api_key="AIzaSyATT0mTSy9xHWOSGzT8IMeL-82xyOm4MSo")
newsdata_client = NewsDataApiClient(apikey="pub_1ab52a22ddb54464bc6c94ac1dd7716e")
INPUT_FILE = "stocks.xlsx"
OUTPUT_FILE = "scores.xlsx"
MAX_WORKERS = 1
LOOKBACK_REDDIT_HOURS  = 24
LOOKBACK_NEWS_HOURS = 24
NEWSAPI_KEY = "4a3dd684d2554c1a9fd7043bd5f859d1"  # <-- Replace with your actual NewsAPI key
WEIGHTS = {
    "google": 0.15,
    "price": 0.25,
    "volume": 0.15,
    "news": 0.45
}
# ==========================

pytrends = TrendReq(hl="en-US", tz=0)
analyzer = SentimentIntensityAnalyzer()
newsapi = NewsApiClient(api_key=NEWSAPI_KEY) if (HAS_NEWSAPI and NEWSAPI_KEY) else None

# -------- Helpers --------
def read_tickers(file_path):
    if not os.path.exists(file_path):
        print(f"❌ Input file not found: {file_path}")
        return []
    df = pd.read_excel(file_path)
    return df.iloc[:, 0].astype(str).str.strip().tolist()

def normalize(value, min_val, max_val):
    if value is None:
        return 0.5
    value = max(min_val, min(value, max_val))
    return (value - min_val) / (max_val - min_val)

# -------- Data Collectors --------
def get_yfinance_signals(ticker):
    try:
        t = yf.Ticker(ticker)
        hist = t.history(period="60d", interval="1d")
        if hist.empty:
            return None, None

        prev_close = hist["Close"].iloc[-1]
        avg_vol = hist["Volume"].tail(50).mean()

        intraday = t.history(period="1d", interval="1m")
        current_price = intraday["Close"].iloc[-1] if not intraday.empty else prev_close
        current_vol = intraday["Volume"].sum() if not intraday.empty else avg_vol

        intraday_pct = (current_price - prev_close) / prev_close
        vol_ratio = current_vol / avg_vol if avg_vol > 0 else 1.0

        return intraday_pct, vol_ratio
    except Exception:
        return None, None


def get_google_trends(ticker):
    kw = ticker.split(".")[0].upper()
    tries = 1
    for attempt in range(tries):
        try:
            pytrends.build_payload([kw], timeframe="now 7-d")
            df = pytrends.interest_over_time()
            if df.empty:
                pytrends.build_payload([kw], timeframe="today 3-m")
                df = pytrends.interest_over_time()
            if not df.empty:
                return float(df[kw].mean())
            return 0.0
        except Exception as e:
            if "429" in str(e):
                print(f"⚠️ Rate limited on {ticker}, waiting 5s before retry...")
                time.sleep(5)  # cooldown
            else:
                print(f"⚠️ Google Trends error for {ticker}: {e}")
                return 0.0
    return 0.0


def get_combined_news_sentiment_gemini(ticker, hours=24, max_articles=20):
    model = genai.GenerativeModel("gemini-1.5-flash")  # Create model outside try

    try:
        articles = []
        to_date = datetime.utcnow()
        from_date = to_date - timedelta(hours=hours)

        # ---- NewsAPI ----
        if newsapi:
            try:
                api_articles = newsapi.get_everything(
                    q=ticker,
                    from_param=from_date.isoformat(),
                    to=to_date.isoformat(),
                    language="en",
                    sort_by="relevancy",
                    page_size=max_articles,
                ).get("articles", [])
                articles.extend(api_articles)
            except Exception as e:
                print(f"⚠️ NewsAPI error for {ticker}: {e}")

        # ---- NewsData.io ----
        try:
            nd_articles = newsdata_client.latest_api(
                q=ticker,
                language="en",
                country="us,in,gb",
                category="business"
            ).get("results", [])
            articles.extend(nd_articles)
        except Exception as e:
            print(f"⚠️ NewsData error for {ticker}: {e}")

        # ---- Combine text ----
        texts = []
        for art in articles:
            title = art.get("title", "")
            desc = art.get("description", "")
            if title or desc:
                texts.append(f"{title}. {desc}")

        if not texts:
            print(f"⚠️ No articles found for {ticker}")
            return None, 0

        # ---- Gemini sentiment ----
        prompt = (
            f"You are a financial analyst. Analyze {len(texts)} recent news snippets "
            f"about {ticker} and rate the overall sentiment between -1 (bearish) and +1 (bullish):\n\n"
            + "\n".join([f"- {t}" for t in texts[:max_articles]])
        )

        response = model.generate_content(prompt)
        text = response.text.strip()

        import re
        match = re.search(r"[-+]?\d*\.\d+|[-+]?\d+", text)
        score = float(match.group()) if match else 0.0
        score = max(-1.0, min(1.0, score))

        return score, len(texts)

    except Exception as e:
        print(f"⚠️ Gemini sentiment error for {ticker}: {e}")
        return None, 0

# -------- Score Calculation --------
def compute_score(ticker):
    print(f"🧩 Processing {ticker} ...")  # Debug log

    # --- Step 1: Get signals ---
    intraday_pct, vol_ratio = get_yfinance_signals(ticker)
    trends = get_google_trends(ticker)
    news_sent, n_articles = get_combined_news_sentiment_gemini(ticker)

    # --- Step 2: Handle None / missing data ---
    intraday_pct = intraday_pct if intraday_pct is not None else 0.0
    vol_ratio = vol_ratio if vol_ratio is not None else 1.0
    trends = trends if trends is not None else 0.0
    news_sent = max(-1.0, min(1.0, news_sent if news_sent is not None else 0.0))

    # --- Step 3: Normalize each metric ---
    price_score = normalize(intraday_pct, -0.1, 0.1)
    volume_score = normalize(vol_ratio, 0, 5)
    google_score = normalize(trends, 0, 100)
    news_score = (news_sent + 1) / 2 if news_sent is not None else 0.5  # convert -1→0, +1→1

    # --- Step 4: Weighted final score ---
    final = (
        WEIGHTS["price"] * price_score +
        WEIGHTS["volume"] * volume_score +
        WEIGHTS["google"] * google_score +
        WEIGHTS["news"] * news_score
    ) * 100

    # --- Step 5: Determine signal ---
    if final > 65:
        signal, action = "Bullish", "Buy"
    elif final >= 45:
        signal, action = "Neutral", "Hold"
    else:
        signal, action = "Bearish", "Avoid"

    # --- Step 6: Return full result ---
    result = {
        "ticker": ticker,
        "score": round(final, 2),
        "signal": signal,
        "action": action,
        "price_score": round(price_score * 100, 2),
        "volume_score": round(volume_score * 100, 2),
        "google_score": round(google_score * 100, 2),
        "news_score": round(news_score * 100, 2),
        "trends": round(trends, 2),
        "news_articles": n_articles,
        "intraday_pct": round(intraday_pct, 4),
        "vol_ratio": round(vol_ratio, 3),
        "news_sentiment": round(news_sent, 3),
    }

    print(f"✅ Done {ticker}: Score={result['score']} ({signal})")  # Debug log
    return result

# -------- Main Function --------
def main():
    tickers = read_tickers(INPUT_FILE)
    if not tickers:
        return

    print(f"📈 Loaded {len(tickers)} tickers from {INPUT_FILE}")

    tickers = tickers[:50]   # process only first 50 tickers

    results = []
    start_time = time.time()

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(compute_score, t) for t in tickers]
        for i, f in enumerate(concurrent.futures.as_completed(futures), 1):
            try:
                results.append(f.result())
                print(f"✅ Processed {i}/{len(tickers)}")
            except Exception as e:
                print(f"⚠️ Error processing ticker {tickers[i-1]}: {e}")

            #Small pause to reduce API throttling
            time.sleep(0.5)

    df = pd.DataFrame(results).sort_values("score", ascending=False)
    df.to_excel(OUTPUT_FILE, index=False)
    df.to_csv(OUTPUT_FILE.replace(".xlsx", ".csv"), index=False)

    # --- Excel color formatting ---
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    header_row = [cell.value for cell in ws[1]]
    signal_col = header_row.index("signal") + 1
    action_col = header_row.index("action") + 1

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        sig = ws.cell(row=row, column=signal_col).value
        act = ws.cell(row=row, column=action_col).value

        fill = green if act == "Buy" else yellow if act == "Hold" else red
        ws.cell(row=row, column=signal_col).fill = fill
        ws.cell(row=row, column=action_col).fill = fill

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Output saved to: {OUTPUT_FILE}")
    print(f" CSV also saved to: {OUTPUT_FILE.replace('.xlsx', '.csv')}")
    print(f" Finished in {time.time() - start_time:.2f} seconds")

if __name__ == "__main__":
    main()
